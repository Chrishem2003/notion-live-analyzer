"""
Academic Audit & Optimization Engine
======================================
Cryptographic blockchain ledger for tamper-proof forensic auditing,
triple-vector linguistic scoring (plagiarism, AI-content detection),
fluid cadence transformation, and advanced text humanization.

Architecture:
  EnterpriseDataEngine        — SHA-256 blockchain ledger (SQLite)
  ProductionLinguisticProcessor — Statistical profiling, steganography, scoring
  TextHumanizer               — Advanced rewriter to evade pattern scanners
  AuditOrchestrator           — High-level orchestrator tying all components
"""
from __future__ import annotations

import base64
import difflib
import hashlib
import io
import json
import os
import re
import sqlite3
import statistics
import time
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from xml.etree import ElementTree

import numpy as np
import pandas as pd
import streamlit as st

from modules.logging_utils import get_logger

logger = get_logger(__name__)

# ─── Paths ────────────────────────────────────────────────────────────
APP_DIR = Path(__file__).resolve().parent.parent
DB_PATH = APP_DIR / "research_workspace.db"


# ═══════════════════════════════════════════════════════════════════════
# 1. ENTERPRISE DATA ENGINE — Cryptographic Blockchain Ledger
# ═══════════════════════════════════════════════════════════════════════
class EnterpriseDataEngine:
    """
    Immutable, tamper-proof audit trail using SHA-256 blockchain.
    Every text change is recorded as a block linked to the previous hash.
    Any tampering is immediately detectable via chain verification.
    """

    def __init__(self, db_path: str | Path = DB_PATH):
        self.db_path = Path(db_path)
        self._init_architecture()

    def _get_connection(self) -> sqlite3.Connection:
        conn = sqlite3.connect(str(self.db_path))
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA journal_mode=WAL")
        return conn

    def _init_architecture(self):
        """Create the audit_ledger table with blockchain structure."""
        with self._get_connection() as conn:
            conn.executescript("""
                CREATE TABLE IF NOT EXISTS audit_ledger (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    project_id INTEGER NOT NULL DEFAULT 0,
                    session_id TEXT NOT NULL,
                    student_id TEXT NOT NULL,
                    timestamp REAL NOT NULL,
                    event_type TEXT NOT NULL,
                    text_snapshot TEXT,
                    payload_metrics TEXT,
                    previous_hash TEXT NOT NULL,
                    current_hash TEXT NOT NULL,
                    account_status TEXT DEFAULT 'ACTIVE',
                    FOREIGN KEY (project_id) REFERENCES projects(id) ON DELETE CASCADE
                );
                CREATE INDEX IF NOT EXISTS idx_audit_ledger_session
                    ON audit_ledger(session_id);
                CREATE INDEX IF NOT EXISTS idx_audit_ledger_project
                    ON audit_ledger(project_id);
            """)
            conn.commit()

            # Auto-migrate: project_id column if missing
            try:
                conn.execute("ALTER TABLE audit_ledger ADD COLUMN project_id INTEGER DEFAULT 0")
                conn.commit()
            except sqlite3.OperationalError as exc:
                if "duplicate column name" not in str(exc).lower():
                    logger.error("Audit ledger migration failed: %s", exc)
                    raise

    def _get_last_block(self, session_id: str) -> str:
        """Get the current_hash of the most recent block in the chain."""
        with self._get_connection() as conn:
            row = conn.execute(
                "SELECT current_hash FROM audit_ledger WHERE session_id = ? ORDER BY id DESC LIMIT 1",
                (session_id,),
            ).fetchone()
            return row["current_hash"] if row else "GENESIS_ROOT_BLOCK_00000000000000000000"

    def record_node(
        self,
        session_id: str,
        student_id: str,
        event_type: str,
        text_snapshot: str = "",
        payload_metrics: str = "",
        project_id: int = 0,
    ) -> str:
        """
        Record a new block in the blockchain ledger.
        Returns the SHA-256 hash of the new block.
        """
        prev_hash = self._get_last_block(session_id)
        ts = time.time()
        block_content = (
            f"{session_id}-{student_id}-{ts}-{event_type}-"
            f"{text_snapshot}-{payload_metrics}-{prev_hash}"
        )
        current_hash = hashlib.sha256(block_content.encode("utf-8")).hexdigest()

        with self._get_connection() as conn:
            conn.execute(
                """INSERT INTO audit_ledger
                   (project_id, session_id, student_id, timestamp, event_type,
                    text_snapshot, payload_metrics, previous_hash, current_hash)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (
                    project_id,
                    session_id,
                    student_id,
                    ts,
                    event_type,
                    text_snapshot,
                    payload_metrics,
                    prev_hash,
                    current_hash,
                ),
            )
            conn.commit()
        return current_hash

    def fetch_and_verify_chain(self, session_id: str) -> Tuple[bool, Any]:
        """
        Fetch and cryptographically verify the entire blockchain for a session.
        Returns (is_valid, timeline_or_error_message).
        """
        with self._get_connection() as conn:
            rows = conn.execute(
                """SELECT timestamp, event_type, text_snapshot, payload_metrics,
                          previous_hash, current_hash, student_id
                   FROM audit_ledger
                   WHERE session_id = ?
                   ORDER BY id ASC""",
                (session_id,),
            ).fetchall()

        if not rows:
            return False, "No audit records found for this session."

        expected_prev = "GENESIS_ROOT_BLOCK_00000000000000000000"
        reconstructed_timeline = []

        for row in rows:
            ts = row["timestamp"]
            event_type = row["event_type"]
            text_snapshot = row["text_snapshot"]
            payload_metrics = row["payload_metrics"]
            prev_hash = row["previous_hash"]
            curr_hash = row["current_hash"]
            student_id = row["student_id"]

            # Verify chain link
            if prev_hash != expected_prev:
                return False, "CRITICAL FAULT: Cryptographic block sequence manipulation detected!"

            # Recompute hash to detect tampering
            block_content = (
                f"{session_id}-{student_id}-{ts}-{event_type}-"
                f"{text_snapshot}-{payload_metrics}-{prev_hash}"
            )
            recalc = hashlib.sha256(block_content.encode("utf-8")).hexdigest()
            if recalc != curr_hash:
                return False, "CRITICAL FAULT: Internal block payload tampering detected!"

            expected_prev = curr_hash
            reconstructed_timeline.append({
                "timestamp": ts,
                "event_type": event_type,
                "text": text_snapshot,
                "metrics": payload_metrics,
                "student_id": student_id,
            })

        return True, reconstructed_timeline

    def get_admin_metrics(self, project_id: int = 0) -> Tuple[Tuple, List]:
        """Get aggregate audit metrics and all records for a project."""
        with self._get_connection() as conn:
            counts = conn.execute(
                """SELECT COUNT(DISTINCT session_id) as sessions,
                          COUNT(id) as total_records,
                          COUNT(DISTINCT student_id) as students
                   FROM audit_ledger
                   WHERE project_id = ?""",
                (project_id,),
            ).fetchone()
            records = conn.execute(
                """SELECT DISTINCT student_id, session_id, account_status
                   FROM audit_ledger
                   WHERE project_id = ?
                   ORDER BY id DESC""",
                (project_id,),
            ).fetchall()
        return (counts["sessions"], counts["total_records"], counts["students"]), [dict(r) for r in records]

    def update_account_status(self, student_id: str, status: str):
        """Update the account status for all records of a student."""
        with self._get_connection() as conn:
            conn.execute(
                "UPDATE audit_ledger SET account_status = ? WHERE student_id = ?",
                (status, student_id),
            )
            conn.commit()

    def get_session_timeline(self, session_id: str) -> List[Dict]:
        """Get the full timeline of events for a session."""
        with self._get_connection() as conn:
            rows = conn.execute(
                """SELECT timestamp, event_type, text_snapshot, payload_metrics, student_id
                   FROM audit_ledger
                   WHERE session_id = ?
                   ORDER BY id ASC""",
                (session_id,),
            ).fetchall()
        return [dict(r) for r in rows]

    def get_event_types(self, session_id: str) -> List[str]:
        """Get distinct event types for a session."""
        with self._get_connection() as conn:
            rows = conn.execute(
                "SELECT DISTINCT event_type FROM audit_ledger WHERE session_id = ?",
                (session_id,),
            ).fetchall()
        return [r["event_type"] for r in rows]


# ═══════════════════════════════════════════════════════════════════════
# 2. PRODUCTION LINGUISTIC PROCESSOR — Triple-Vector Forensic Scanner
# ═══════════════════════════════════════════════════════════════════════
class ProductionLinguisticProcessor:
    """
    Statistical text analysis and transformation engine.
    Provides:
      - Statistical profiling (burstiness, perplexity, sentence complexity)
      - Fluid cadence transformation (structural rewriting)
      - Steganographic watermark embedding (zero-width characters)
      - Triangulated scoring (plagiarism, AI-content, authenticity)
      - N-gram overlap plagiarism detection
    """

    # Common AI-generated phrases for heuristic detection
    AI_PATTERNS = [
        r"\bAs an AI\b",
        r"\bI don't have personal\b",
        r"\bI cannot provide\b",
        r"\bI'm sorry, but I cannot\b",
        r"\bAs a language model\b",
        r"\bI do not have access to\b",
        r"\bIt is important to note that\b",
        r"\bIn today's rapidly evolving\b",
        r"\bThe landscape of\b",
        r"\bThis comprehensive analysis\b",
        r"\bDelving into\b",
        r"\bIt is worth noting that\b",
        r"\bIn the realm of\b",
        r"\bWhen it comes to\b",
        r"\bA plethora of\b",
        r"\bThe fact of the matter is\b",
        r"\bIn the grand scheme of things\b",
        r"\bIt goes without saying\b",
        r"\bIn this ever-changing world\b",
        r"\bUnprecedented\b",
        r"\bIn the current climate\b",
        r"\bThe very fabric of\b",
        r"\bAt the end of the day\b",
        r"\bIt remains to be seen\b",
        r"\bAll things considered\b",
    ]

    ALTERNATIVES = {
        r"\bFurthermore\b": "In tandem with these observations,",
        r"\bIn conclusion\b": "Ultimately, the structural evidence points to",
        r"\bMoreover\b": "Correspondingly,",
        r"\bTherefore\b": "Consequently,",
        r"\bIt is important to note that\b": "Notably,",
        r"\bAdditionally\b": "Complementing this,",
        r"\bHowever\b": "That said,",
        r"\bConsequently\b": "As a direct result,",
        r"\bNevertheless\b": "Counter to this,",
        r"\bFurthermore\b": "Extending this line of inquiry,",
        r"\bIn addition\b": "Augmenting this perspective,",
        r"\bOn the other hand\b": "Conversely,",
        r"\bIn particular\b": "Specifically,",
        r"\bAs a result\b": "Consequently,",
        r"\bIn contrast\b": "By comparison,",
    }

    @staticmethod
    def run_statistical_profile(text: str) -> Dict[str, float]:
        """
        Compute statistical text metrics:
          - burstiness: std deviation of sentence lengths
          - perplexity: unique word ratio * 100
          - sentences: sentence count
          - avg_sentence_length: average words per sentence
          - vocabulary_richness: type-token ratio
        """
        if not text or not text.strip():
            return {
                "burstiness": 0.0,
                "perplexity": 0.0,
                "sentences": 0,
                "avg_sentence_length": 0.0,
                "vocabulary_richness": 0.0,
                "total_words": 0,
                "unique_words": 0,
            }

        sentences = [s.strip() for s in re.split(r"[.!?]", text) if s.strip()]
        lengths = [len(s.split()) for s in sentences if s.split()]

        words = re.findall(r"\b\w+\b", text.lower())
        unique_words = len(set(words))
        total_words = len(words) if words else 1

        burstiness = round(float(np.std(lengths)), 2) if len(lengths) >= 2 else 0.0
        perplexity = round((unique_words / total_words) * 100, 2)
        avg_sent_len = round(statistics.mean(lengths), 2) if lengths else 0.0
        vocabulary_richness = round(unique_words / total_words, 4) if words else 0.0

        return {
            "burstiness": burstiness,
            "perplexity": perplexity,
            "sentences": len(sentences),
            "avg_sentence_length": avg_sent_len,
            "vocabulary_richness": vocabulary_richness,
            "total_words": total_words,
            "unique_words": unique_words,
        }

    @staticmethod
    def detect_ai_patterns(text: str) -> Dict[str, Any]:
        """
        Scan text for common AI-generated phrasing patterns.
        Returns pattern matches and an AI-likelihood score (0-100).
        """
        matches = []
        for pattern in ProductionLinguisticProcessor.AI_PATTERNS:
            found = re.findall(pattern, text, flags=re.IGNORECASE)
            if found:
                matches.extend(found)

        # Score based on pattern density
        words = len(re.findall(r"\b\w+\b", text))
        if words == 0:
            return {"matches": [], "ai_pattern_score": 0.0, "pattern_count": 0}

        pattern_count = len(matches)
        # Normalize: more patterns per 1000 words = higher AI likelihood
        density = (pattern_count / max(words, 1)) * 1000
        ai_pattern_score = min(100.0, round(density * 25, 2))  # Scale

        return {
            "matches": matches[:20],  # Limit to 20 for display
            "ai_pattern_score": ai_pattern_score,
            "pattern_count": pattern_count,
        }

    @staticmethod
    def process_fluid_cadence(text: str) -> str:
        """
        Structural text transformation that rewrites sentences to
        vary cadence and rhythm. Replaces common transitional phrases
        and restructures sentence mid-points.
        """
        if not text or not text.strip():
            return text

        sentences = [s.strip() for s in re.split(r"[.!?]", text) if s.strip()]
        transformed = []

        for idx, sent in enumerate(sentences):
            if not sent.strip():
                continue

            # Apply alternative phrasing substitutions
            for pat, rep in ProductionLinguisticProcessor.ALTERNATIVES.items():
                sent = re.sub(pat, rep, sent, flags=re.IGNORECASE)

            words = sent.split()
            if not words:
                continue

            # Every 3rd sentence: mid-point restructuring (if > 14 words)
            if idx % 3 == 0 and len(words) > 14:
                mid = len(words) // 2
                first_half = " ".join(words[:mid])
                second_half = " ".join(words[mid:]).lower()
                transformed.append(
                    f"{first_half}; this directly informs why {second_half}"
                )
            # Every 4th sentence: front-load with "Crucially" (if short)
            elif idx % 4 == 0 and len(words) < 8 and len(transformed) > 0:
                # Only transform if not already at start
                lower_text = " ".join(words).lower()
                transformed.append(f"Crucially, {lower_text}")
            else:
                transformed.append(" ".join(words))

        return ". ".join(transformed) + "."

    @staticmethod
    def embed_steganography(text: str, marker: str = "") -> str:
        """
        Embed invisible zero-width characters as steganographic watermarks.
        Every 3rd word gets a zero-width space appended.
        Can optionally prepend a marker string.
        """
        if not text or not text.strip():
            return text

        # Zero-width characters
        zero_width_space = "\u200b"
        zero_width_joiner = "\u200d"
        zero_width_non_joiner = "\u200c"

        # Embed marker at start if provided
        if marker:
            marker_encoded = "".join(
                f"{char}{zero_width_space}"
                for char in hashlib.md5(marker.encode()).hexdigest()[:8]
            )
            result = marker_encoded + " "
        else:
            result = ""

        # Add watermark to every 3rd word and random others
        words = text.split(" ")
        for idx, word in enumerate(words):
            if idx % 3 == 0:
                result += word + zero_width_space + " "
            elif idx % 7 == 0:
                result += word + zero_width_joiner + " "
            elif idx % 11 == 0:
                result += word + zero_width_non_joiner + " "
            else:
                result += word + " "

        return result.strip()

    @staticmethod
    def clean_for_export(text: str) -> str:
        """Strip all zero-width characters and steganographic markers."""
        if not text:
            return text
        # Remove zero-width spaces, joiners, non-joiners
        cleaned = text.replace("\u200b", "").replace("\u200c", "").replace("\u200d", "")
        # Remove any other invisible Unicode control chars
        cleaned = re.sub(r"[\u200e\u200f\u2028\u2029\u202a\u202b\u202c\u202d\u202e\u2060\u2061\u2062\u2063\u2064]", "", cleaned)
        return cleaned.strip()

    @staticmethod
    def calculate_triangulated_scores(trail: List[Dict]) -> Dict[str, Any]:
        """
        Calculate multi-vector scores from an audit trail:
          - plagiarism: estimated plagiarism risk (0-100)
          - ai_content: estimated AI-generated content score (0-100)
          - authenticity: human authorship authenticity score (0-100)
          - time_delta: time span of the trail in seconds
          - total_events: number of events in the trail
        """
        total_nodes = len(trail)
        if total_nodes == 0:
            return {
                "plagiarism": 100.0,
                "ai_content": 100.0,
                "authenticity": 0.0,
                "time_delta": 0,
                "total_events": 0,
                "import_ratio": 0.0,
                "conversion_ratio": 0.0,
                "labor_ratio": 0.0,
                "aidify_score": 0.0,
            }

        # Count anomaly types
        imports = sum(1 for n in trail if n["event_type"] == "MASS_EXTERNAL_IMPORT_ANOMALY")
        conversions = sum(1 for n in trail if n["event_type"] == "CADENCE_CONVERSION_APPLIED")
        ai_detections = sum(
            1 for n in trail
            if n["event_type"] in ("AI_CONTENT_DETECTED", "HIGH_AI_PROBABILITY")
        )
        paste_events = sum(
            1 for n in trail
            if n["event_type"] in ("BULK_PASTE_DETECTED", "EXTERNAL_PASTE")
        )

        # Time span analysis
        timestamps = [n["timestamp"] for n in trail if isinstance(n.get("timestamp"), (int, float))]
        time_span = round(max(timestamps) - min(timestamps), 1) if len(timestamps) > 1 else 0

        # Plagiarism score: based on import/paste anomalies
        import_ratio = imports / max(total_nodes, 1)
        paste_ratio = paste_events / max(total_nodes, 1)
        plagiarism = round(min(100.0, (import_ratio * 70 + paste_ratio * 30) * 100), 1)

        # AI content score: based on AI detections + conversion events
        ai_ratio = ai_detections / max(total_nodes, 1)
        conversion_ratio = conversions / max(total_nodes, 1)
        ai_content = round(min(100.0, (ai_ratio * 60 + conversion_ratio * 40) * 100), 1)

        # Authenticity: inverse of plagiarism + AI, with labor ratio boost
        labor_ratio = (total_nodes - imports - paste_events) / max(total_nodes, 1)
        aidify_score = round(labor_ratio * 100, 1)
        authenticity = round(max(0.0, 100.0 - (plagiarism * 0.5 + ai_content * 0.3)), 1)

        return {
            "plagiarism": plagiarism,
            "ai_content": ai_content,
            "authenticity": authenticity,
            "time_delta": time_span,
            "total_events": total_nodes,
            "import_ratio": round(import_ratio * 100, 1),
            "conversion_ratio": round(conversion_ratio * 100, 1),
            "labor_ratio": round(labor_ratio * 100, 1),
            "aidify_score": aidify_score,
        }

    @staticmethod
    def ngram_plagiarism_check(
        text: str,
        reference_texts: List[str],
        n: int = 5,
    ) -> Dict[str, Any]:
        """
        Check text for n-gram overlap with reference texts.
        Returns overlap statistics and similarity scores.
        """
        if not text or not reference_texts:
            return {"overall_similarity": 0.0, "ngram_matches": [], "matched_count": 0}

        def get_ngrams(t: str, size: int) -> set:
            words = re.findall(r"\b\w+\b", t.lower())
            return set(" ".join(words[i : i + size]) for i in range(len(words) - size + 1))

        text_ngrams = get_ngrams(text, n)
        if not text_ngrams:
            return {"overall_similarity": 0.0, "ngram_matches": [], "matched_count": 0}

        all_reference_ngrams = set()
        for ref in reference_texts:
            all_reference_ngrams |= get_ngrams(ref, n)

        if not all_reference_ngrams:
            return {"overall_similarity": 0.0, "ngram_matches": [], "matched_count": 0}

        matches = text_ngrams & all_reference_ngrams
        similarity = round(len(matches) / len(text_ngrams) * 100, 2)

        return {
            "overall_similarity": similarity,
            "ngram_matches": sorted(matches)[:50],  # Limit for display
            "matched_count": len(matches),
            "total_ngrams": len(text_ngrams),
        }


# ═══════════════════════════════════════════════════════════════════════
# 3. TEXT HUMANIZER — Advanced AI-Evasion Rewriting Framework
# ═══════════════════════════════════════════════════════════════════════
class TextHumanizer:
    """
    Advanced text rewriting engine designed to programmatically
    randomize burstiness and perplexity to evade pattern scanners.
    """

    # Synonym maps for common academic words
    SYNONYM_MAP = {
        "important": ["significant", "crucial", "essential", "vital", "critical", "paramount"],
        "shows": ["demonstrates", "indicates", "reveals", "suggests", "illustrates", "highlights"],
        "therefore": ["consequently", "thus", "hence", "accordingly", "as a result", "for this reason"],
        "however": ["nevertheless", "nonetheless", "that said", "conversely", "on the other hand"],
        "many": ["numerous", "multiple", "various", "countless", "a multitude of", "several"],
        "big": ["substantial", "considerable", "significant", "sizeable", "extensive", "large-scale"],
        "small": ["minor", "limited", "modest", "marginal", "incremental", "diminutive"],
        "good": ["positive", "favorable", "beneficial", "advantageous", "constructive"],
        "bad": ["negative", "adverse", "detrimental", "unfavorable", "deleterious"],
        "different": ["distinct", "divergent", "varying", "disparate", "contrasting", "dissimilar"],
        "useful": ["valuable", "beneficial", "effective", "practical", "instrumental", "advantageous"],
        "change": ["transform", "modify", "alter", "adjust", "adapt", "refine"],
        "result": ["outcome", "consequence", "effect", "finding", "implication"],
        "study": ["investigation", "analysis", "examination", "inquiry", "exploration", "research"],
        "method": ["approach", "technique", "methodology", "procedure", "protocol"],
        "example": ["instance", "illustration", "case", "demonstration", "sample"],
        "focus": ["emphasis", "concentration", "central theme", "core aspect", "primary concern"],
        "process": ["procedure", "mechanism", "workflow", "method", "pipeline", "framework"],
        "increase": ["rise", "growth", "expansion", "elevation", "upsurge", "escalation"],
        "decrease": ["reduction", "decline", "diminution", "downturn", "drop", "attenuation"],
    }

    @staticmethod
    def randomize_burstiness(text: str, target_burstiness: float = 3.5) -> str:
        """
        Restructure sentences to achieve a target burstiness (std dev of
        sentence lengths). Higher burstiness = more natural human writing.
        """
        if not text or not text.strip():
            return text

        sentences = [s.strip() for s in re.split(r"[.!?]", text) if s.strip()]
        if len(sentences) < 3:
            return text

        current_lengths = [len(s.split()) for s in sentences if s.split()]
        if not current_lengths:
            return text

        current_burstiness = float(np.std(current_lengths)) if len(current_lengths) > 1 else 0

        # Only adjust if current burstiness is outside target range
        if abs(current_burstiness - target_burstiness) < 0.5:
            return text

        # Strategy: merge/split sentences to adjust burstiness
        adjusted = []
        i = 0
        while i < len(sentences):
            words = sentences[i].split()
            word_count = len(words)

            if word_count < 5 and i + 1 < len(sentences):
                # Merge with next sentence for longer block
                next_words = sentences[i + 1].split()
                merged = words + ["and"] + next_words
                adjusted.append(" ".join(merged))
                i += 2
            elif word_count > 25 and ";" not in sentences[i]:
                # Split long sentence at conjunction or mid-point
                mid = word_count // 2
                first = " ".join(words[:mid])
                second = " ".join(words[mid:])
                adjusted.append(first + ",")
                adjusted.append(second.lower())
                i += 1
            else:
                adjusted.append(sentences[i])
                i += 1

        return ". ".join(adjusted) + "."

    @staticmethod
    def randomize_perplexity(text: str, target_perplexity: float = 65.0) -> str:
        """
        Adjust vocabulary uniqueness to achieve target perplexity.
        Higher perplexity = more diverse vocabulary = more human-like.
        """
        if not text or not text.strip():
            return text

        words = re.findall(r"\b\w+\b", text.lower())
        if not words:
            return text

        unique = set(words)
        current_perplexity = (len(unique) / len(words)) * 100

        if abs(current_perplexity - target_perplexity) < 5:
            return text

        # Apply synonym substitution to shift perplexity
        result_words = []
        for word in words:
            if word in TextHumanizer.SYNONYM_MAP and np.random.random() < 0.3:
                # Substitute with a synonym
                synonyms = TextHumanizer.SYNONYM_MAP[word]
                result_words.append(np.random.choice(synonyms))
            else:
                result_words.append(word)

        return " ".join(result_words)

    @staticmethod
    def structural_rewrite(text: str) -> str:
        """
        Apply structural transformations to vary sentence patterns:
        - Front-loading adverbial phrases
        - Mid-sentence restructuring
        - Varying sentence openings
        """
        if not text or not text.strip():
            return text

        sentences = [s.strip() for s in re.split(r"[.!?]", text) if s.strip()]
        transformed = []

        openers = [
            "Notably,", "Interestingly,", "Significantly,",
            "Critically,", "Importantly,", "Strikingly,",
            "Of particular interest,", "Noteworthy is that",
            "A key observation is that", "Remarkably,",
            "Fundamentally,", "Essentially,", "Crucially,",
        ]

        for idx, sent in enumerate(sentences):
            if not sent.strip():
                continue

            words = sent.split()
            if not words:
                continue

            # Every 5th sentence: add introductory phrase
            if idx % 5 == 0 and idx > 0 and len(words) > 6:
                opener = np.random.choice(openers)
                transformed.append(f"{opener} {' '.join(words).lower()}")
            # Every 7th sentence: reposition temporal clause
            elif idx % 7 == 0 and len(words) > 10:
                mid = len(words) // 3
                clause = " ".join(words[:mid])
                rest = " ".join(words[mid:])
                transformed.append(f"After {clause}, {rest.lower()}")
            else:
                transformed.append(" ".join(words))

        return ". ".join(transformed) + "."

    @staticmethod
    def full_humanize_pipeline(
        text: str,
        target_burstiness: float = 3.5,
        target_perplexity: float = 65.0,
        apply_cadence: bool = True,
        apply_structural: bool = True,
    ) -> str:
        """
        Complete humanization pipeline:
        1. Clean text
        2. Fluid cadence transformation (optional)
        3. Burstiness randomization
        4. Perplexity randomization
        5. Structural rewrite (optional)
        """
        if not text or not text.strip():
            return text

        result = text

        # Step 1: Clean
        result = ProductionLinguisticProcessor.clean_for_export(result)

        # Step 2: Fluid cadence
        if apply_cadence:
            result = ProductionLinguisticProcessor.process_fluid_cadence(result)

        # Step 3: Burstiness randomization
        result = TextHumanizer.randomize_burstiness(result, target_burstiness)

        # Step 4: Perplexity randomization
        result = TextHumanizer.randomize_perplexity(result, target_perplexity)

        # Step 5: Structural rewrite
        if apply_structural:
            result = TextHumanizer.structural_rewrite(result)

        return result


# ═══════════════════════════════════════════════════════════════════════
# 4. UNIVERSAL FILE READER — Extract text from any document format
# ═══════════════════════════════════════════════════════════════════════
class UniversalFileReader:
    """
    Extract text content from virtually any file format.
    Supports: TXT, DOCX, PDF, CSV, MD, HTML, RTF, ODT, XLSX, IPYNB,
              JSON, XML, PY, R, CPP, JS, TS, and more.
    """

    SUPPORTED_EXTENSIONS = {
        ".txt", ".docx", ".pdf", ".csv", ".md", ".html", ".htm",
        ".rtf", ".odt", ".xlsx", ".xls", ".ipynb", ".json", ".xml",
        ".py", ".r", ".cpp", ".c", ".h", ".hpp", ".js", ".ts", ".java",
        ".rs", ".go", ".rb", ".php", ".swift", ".kt", ".scala", ".sh",
        ".yaml", ".yml", ".toml", ".ini", ".cfg", ".log",
    }

    @classmethod
    def read_file(cls, file_bytes: bytes, filename: str) -> Tuple[str, str]:
        """
        Read text content from a file.
        Returns (extracted_text, error_message).
        If error_message is non-empty, extracted_text will be empty.
        """
        ext = Path(filename).suffix.lower()

        try:
            if ext == ".txt":
                return cls._read_txt(file_bytes), ""
            elif ext == ".docx":
                return cls._read_docx(file_bytes), ""
            elif ext == ".pdf":
                return cls._read_pdf(file_bytes), ""
            elif ext == ".csv":
                return cls._read_csv(file_bytes), ""
            elif ext == ".md":
                return cls._read_txt(file_bytes), ""
            elif ext in (".html", ".htm"):
                return cls._read_html(file_bytes), ""
            elif ext == ".rtf":
                return cls._read_txt(file_bytes), ""
            elif ext == ".odt":
                return cls._read_odt(file_bytes), ""
            elif ext in (".xlsx", ".xls"):
                return cls._read_xlsx(file_bytes), ""
            elif ext == ".ipynb":
                return cls._read_ipynb(file_bytes), ""
            elif ext == ".json":
                return cls._read_json(file_bytes), ""
            elif ext == ".xml":
                return cls._read_xml(file_bytes), ""
            elif ext in (".py", ".r", ".cpp", ".c", ".h", ".hpp", ".js", ".ts",
                         ".java", ".rs", ".go", ".rb", ".php", ".swift", ".kt",
                         ".scala", ".sh", ".yaml", ".yml", ".toml", ".ini",
                         ".cfg", ".log"):
                return cls._read_code_file(file_bytes, ext), ""
            else:
                # Try as plain text
                try:
                    return file_bytes.decode("utf-8"), ""
                except UnicodeDecodeError:
                    return "", f"Unsupported file format: {ext}. Please convert to TXT or DOCX."
        except Exception as e:
            return "", f"Error reading file: {str(e)}"

    @classmethod
    def _read_txt(cls, data: bytes) -> str:
        """Read plain text file."""
        try:
            return data.decode("utf-8")
        except UnicodeDecodeError:
            try:
                return data.decode("latin-1")
            except Exception:
                return data.decode("utf-8", errors="replace")

    @classmethod
    def _read_docx(cls, data: bytes) -> str:
        """Read text from DOCX file."""
        try:
            from docx import Document as DocxDocument
            doc = DocxDocument(io.BytesIO(data))
            paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
            return "\n".join(paragraphs)
        except ImportError:
            # Fallback: try to extract XML content
            try:
                import zipfile
                with zipfile.ZipFile(io.BytesIO(data)) as z:
                    xml_content = z.read("word/document.xml")
                    root = ElementTree.fromstring(xml_content)
                    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
                    texts = []
                    for t in root.iter("{http://schemas.openxmlformats.org/wordprocessingml/2006/main}t"):
                        if t.text:
                            texts.append(t.text)
                    return " ".join(texts)
            except Exception:
                return "DOCX parsing requires python-docx. Install: pip install python-docx"

    @classmethod
    def _read_pdf(cls, data: bytes) -> str:
        """Extract text from PDF file."""
        try:
            import PyPDF2
            reader = PyPDF2.PdfReader(io.BytesIO(data))
            texts = [page.extract_text() for page in reader.pages if page.extract_text()]
            return "\n".join(texts)
        except ImportError:
            try:
                import pdfplumber
                with pdfplumber.open(io.BytesIO(data)) as pdf:
                    texts = [page.extract_text() for page in pdf.pages if page.extract_text()]
                    return "\n".join(texts)
            except ImportError:
                try:
                    import pdfminer
                    from pdfminer.high_level import extract_text
                    return extract_text(io.BytesIO(data))
                except ImportError:
                    return "PDF parsing requires PyPDF2 or pdfplumber. Install: pip install PyPDF2"

    @classmethod
    def _read_csv(cls, data: bytes) -> str:
        """Read CSV file as text."""
        try:
            text = data.decode("utf-8")
        except UnicodeDecodeError:
            text = data.decode("latin-1")
        # Return as tabular text representation
        try:
            df = pd.read_csv(io.StringIO(text))
            return df.to_string(index=False)
        except Exception:
            return text[:100000]  # Limit size

    @classmethod
    def _read_html(cls, data: bytes) -> str:
        """Extract text from HTML file."""
        try:
            text = data.decode("utf-8")
        except UnicodeDecodeError:
            text = data.decode("latin-1")
        # Strip HTML tags
        clean = re.sub(r"<[^>]+>", " ", text)
        clean = re.sub(r"\s+", " ", clean).strip()
        return clean[:100000]  # Limit size

    @classmethod
    def _read_odt(cls, data: bytes) -> str:
        """Extract text from ODT file."""
        try:
            import zipfile
            with zipfile.ZipFile(io.BytesIO(data)) as z:
                content = z.read("content.xml")
                root = ElementTree.fromstring(content)
                ns = {
                    "text": "urn:oasis:names:tc:opendocument:xmlns:text:1.0",
                    "office": "urn:oasis:names:tc:opendocument:xmlns:office:1.0",
                }
                texts = []
                for elem in root.iter():
                    if elem.tag.endswith("}p") or elem.tag.endswith("}h"):
                        if elem.text:
                            texts.append(elem.text)
                    if elem.tag.endswith("}span") and elem.text:
                        texts.append(elem.text)
                return " ".join(texts)
        except Exception:
            return "ODT parsing error. Ensure the file is valid."

    @classmethod
    def _read_xlsx(cls, data: bytes) -> str:
        """Read Excel file as text."""
        try:
            df = pd.read_excel(io.BytesIO(data), sheet_name=None)
            parts = []
            for sheet_name, sheet_df in df.items():
                parts.append(f"--- Sheet: {sheet_name} ---")
                parts.append(sheet_df.to_string(index=False))
            return "\n".join(parts)
        except Exception:
            try:
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True)
                texts = []
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    for row in ws.iter_rows(values_only=True):
                        row_text = " | ".join(str(c) for c in row if c is not None)
                        if row_text.strip():
                            texts.append(row_text)
                return "\n".join(texts)
            except Exception as e:
                return f"XLSX parsing error: {str(e)}"

    @classmethod
    def _read_ipynb(cls, data: bytes) -> str:
        """Extract text from Jupyter notebook file."""
        try:
            notebook = json.loads(data.decode("utf-8"))
            texts = []
            for cell in notebook.get("cells", []):
                source = cell.get("source", [])
                if isinstance(source, list):
                    source = "".join(source)
                texts.append(source)
            return "\n\n".join(texts)
        except Exception as e:
            return f"IPYNB parsing error: {str(e)}"

    @classmethod
    def _read_json(cls, data: bytes) -> str:
        """Read JSON file as formatted text."""
        try:
            obj = json.loads(data.decode("utf-8"))
            return json.dumps(obj, indent=2)
        except Exception as e:
            return f"JSON parsing error: {str(e)}"

    @classmethod
    def _read_xml(cls, data: bytes) -> str:
        """Read XML file as text."""
        try:
            root = ElementTree.fromstring(data)
            return ElementTree.tostring(root, encoding="unicode")
        except Exception as e:
            return f"XML parsing error: {str(e)}"

    @classmethod
    def _read_code_file(cls, data: bytes, ext: str) -> str:
        """Read source code file."""
        try:
            return data.decode("utf-8")
        except UnicodeDecodeError:
            try:
                return data.decode("latin-1")
            except Exception:
                return data.decode("utf-8", errors="replace")


# ═══════════════════════════════════════════════════════════════════════
# 5. AUDIT ORCHESTRATOR — High-Level Coordination Layer
# ═══════════════════════════════════════════════════════════════════════
class AuditOrchestrator:
    """
    Coordinates the full audit workflow:
      1. Initialize EnterpriseDataEngine for blockchain ledger
      2. Use ProductionLinguisticProcessor for scoring
      3. Use TextHumanizer for optimization
      4. Use UniversalFileReader for multi-format file handling
      5. Integrate with LiteratureDatabase for report sections
    """

    def __init__(self, project_id: int = 0):
        self.project_id = project_id
        self.ledger = EnterpriseDataEngine()
        self.processor = ProductionLinguisticProcessor()
        self.humanizer = TextHumanizer()
        self.reader = UniversalFileReader()
        self._session_id = f"audit_{int(time.time())}_{project_id}"

    @property
    def session_id(self) -> str:
        return self._session_id

    def audit_text(
        self,
        text: str,
        student_id: str = "anonymous",
        reference_texts: Optional[List[str]] = None,
    ) -> Dict[str, Any]:
        """
        Perform a complete audit on a text:
          1. Statistical profiling
          2. AI pattern detection
          3. N-gram plagiarism check (if references provided)
          4. Record in blockchain ledger
          5. Return comprehensive results
        """
        if not text or not text.strip():
            return {"error": "No text provided for audit."}

        results = {}

        # Step 1: Statistical profile
        profile = self.processor.run_statistical_profile(text)
        results["statistical_profile"] = profile

        # Step 2: AI pattern detection
        ai_results = self.processor.detect_ai_patterns(text)
        results["ai_detection"] = ai_results

        # Step 3: Plagiarism check if references provided
        if reference_texts:
            plagiarism = self.processor.ngram_plagiarism_check(text, reference_texts)
            results["plagiarism_check"] = plagiarism
        else:
            results["plagiarism_check"] = None

        # Step 4: Compute composite scores
        # From the linguistic profile alone, estimate scores
        words = profile.get("total_words", 0)
        sentences = profile.get("sentences", 0)

        # Perplexity-based AI score (very low or very high perplexity = suspicious)
        perplexity = profile.get("perplexity", 50)
        if perplexity > 85:
            ai_likelihood = 70 + (perplexity - 85) * 2  # 70-100
        elif perplexity < 15:
            ai_likelihood = 60 + (15 - perplexity) * 2  # 60-90
        else:
            ai_likelihood = 50 - abs(perplexity - 50) * 0.8  # 22-50

        ai_content_score = max(0, min(100, round(ai_likelihood + ai_results["ai_pattern_score"] * 0.3, 1)))
        authenticity_score = max(0, min(100, round(100 - ai_content_score * 0.6, 1)))
        plagiarism_score = results["plagiarism_check"]["overall_similarity"] if results["plagiarism_check"] else 0.0

        results["composite_scores"] = {
            "ai_content_score": ai_content_score,
            "authenticity_score": authenticity_score,
            "plagiarism_score": plagiarism_score,
            "overall_risk": round((ai_content_score + plagiarism_score) / 2, 1),
        }

        # Step 5: Record in blockchain ledger
        payload = json.dumps({
            "words": words,
            "sentences": sentences,
            "ai_score": ai_content_score,
            "authenticity": authenticity_score,
        })
        self.ledger.record_node(
            session_id=self._session_id,
            student_id=student_id,
            event_type="TEXT_AUDIT_COMPLETED",
            text_snapshot=text[:500],  # Store preview
            payload_metrics=payload,
            project_id=self.project_id,
        )

        return results

    def audit_report_sections(
        self,
        sections: List[Dict],
        bibliography: Optional[List[Dict]] = None,
        student_id: str = "researcher",
    ) -> List[Dict]:
        """
        Audit all report sections for a project.
        Compares against bibliography texts for plagiarism.
        """
        if not sections:
            return []

        # Extract reference texts from bibliography
        reference_texts = []
        if bibliography:
            for paper in bibliography:
                title = paper.get("title", "")
                abstract = paper.get("abstract", "")
                notes = paper.get("user_notes", "")
                findings = paper.get("user_findings", "")
                combined = f"{title} {abstract} {notes} {findings}".strip()
                if len(combined) > 50:
                    reference_texts.append(combined)

        audit_results = []
        for section in sections:
            content = section.get("content", "").strip()
            if not content:
                continue

            result = self.audit_text(
                text=content,
                student_id=student_id,
                reference_texts=reference_texts,
            )
            result["section_title"] = section.get("section_title", "Untitled")
            result["section_id"] = section.get("id", 0)
            audit_results.append(result)

        return audit_results

    def optimize_text(
        self,
        text: str,
        mode: str = "light",
        target_burstiness: float = 3.5,
        target_perplexity: float = 65.0,
    ) -> Dict[str, Any]:
        """
        Optimize/humanize text with selected mode:
          - "light": Fluid cadence only
          - "balanced": Cadence + burstiness adjustment
          - "deep": Full pipeline (cadence + burstiness + perplexity + structural)
        """
        if not text or not text.strip():
            return {"error": "No text provided for optimization."}

        original_stats = self.processor.run_statistical_profile(text)

        if mode == "light":
            optimized = self.processor.process_fluid_cadence(text)
        elif mode == "balanced":
            optimized = self.humanizer.full_humanize_pipeline(
                text,
                target_burstiness=target_burstiness,
                target_perplexity=target_perplexity,
                apply_cadence=True,
                apply_structural=False,
            )
        else:  # deep
            optimized = self.humanizer.full_humanize_pipeline(
                text,
                target_burstiness=target_burstiness,
                target_perplexity=target_perplexity,
                apply_cadence=True,
                apply_structural=True,
            )

        optimized_stats = self.processor.run_statistical_profile(optimized)

        # Record in ledger
        self.ledger.record_node(
            session_id=self._session_id,
            student_id="optimizer",
            event_type=f"CADENCE_CONVERSION_APPLIED_{mode.upper()}",
            text_snapshot=optimized[:500],
            payload_metrics=json.dumps({
                "original_burstiness": original_stats["burstiness"],
                "new_burstiness": optimized_stats["burstiness"],
                "original_perplexity": original_stats["perplexity"],
                "new_perplexity": optimized_stats["perplexity"],
            }),
            project_id=self.project_id,
        )

        return {
            "original_text": text,
            "optimized_text": optimized,
            "original_stats": original_stats,
            "optimized_stats": optimized_stats,
            "mode": mode,
            "changes": {
                "burstiness_delta": round(optimized_stats["burstiness"] - original_stats["burstiness"], 2),
                "perplexity_delta": round(optimized_stats["perplexity"] - original_stats["perplexity"], 2),
                "sentence_delta": optimized_stats["sentences"] - original_stats["sentences"],
            },
        }

    def create_forensic_timeline(self, start_text: str, end_text: str) -> List[Dict]:
        """
        Generate a forensic diff timeline between two text versions.
        Returns word-by-word changes with timestamps.
        """
        start_words = start_text.split()
        end_words = end_text.split()

        diff = list(difflib.ndiff(start_words, end_words))
        timeline = []
        ts = time.time()

        for i, change in enumerate(diff):
            if change.startswith("  "):  # Unchanged
                continue
            elif change.startswith("- "):  # Removed
                timeline.append({
                    "timestamp": ts + i * 0.1,
                    "event_type": "WORD_REMOVED",
                    "text": change[2:],
                })
            elif change.startswith("+ "):  # Added
                timeline.append({
                    "timestamp": ts + i * 0.1,
                    "event_type": "WORD_ADDED",
                    "text": change[2:],
                })
            elif change.startswith("? "):  # Change indicator
                continue

        return timeline

    def generate_export_report(self, audit_results: List[Dict]) -> str:
        """Generate a formatted audit report string."""
        lines = [
            "═" * 70,
            "ACADEMIC AUDIT & COMPLIANCE REPORT",
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            f"Session ID: {self._session_id}",
            "═" * 70,
            "",
        ]

        for result in audit_results:
            if "error" in result:
                lines.append(f"ERROR: {result['error']}")
                continue

            section_title = result.get("section_title", "Unknown Section")
            lines.append(f"Section: {section_title}")
            lines.append("-" * 40)

            stats = result.get("statistical_profile", {})
            lines.append(f"  Words: {stats.get('total_words', 'N/A')}")
            lines.append(f"  Sentences: {stats.get('sentences', 'N/A')}")
            lines.append(f"  Burstiness: {stats.get('burstiness', 'N/A')}")
            lines.append(f"  Perplexity: {stats.get('perplexity', 'N/A')}%")
            lines.append(f"  Vocabulary Richness: {stats.get('vocabulary_richness', 'N/A')}")

            scores = result.get("composite_scores", {})
            lines.append(f"  AI Content Score: {scores.get('ai_content_score', 'N/A')}%")
            lines.append(f"  Authenticity Score: {scores.get('authenticity_score', 'N/A')}%")
            lines.append(f"  Overall Risk: {scores.get('overall_risk', 'N/A')}%")

            ai_det = result.get("ai_detection", {})
            lines.append(f"  AI Pattern Matches: {ai_det.get('pattern_count', 'N/A')}")

            plag = result.get("plagiarism_check", {})
            if plag:
                lines.append(f"  N-Gram Similarity: {plag.get('overall_similarity', 'N/A')}%")
                lines.append(f"  Matching N-Grams: {plag.get('matched_count', 'N/A')}")

            lines.append("")

        lines.append("═" * 70)
        lines.append("END OF AUDIT REPORT")
        lines.append("═" * 70)

        return "\n".join(lines)


# ═══════════════════════════════════════════════════════════════════════
# Module-level singleton
# ═══════════════════════════════════════════════════════════════════════
_audit_orchestrator: Optional[AuditOrchestrator] = None


def get_audit_orchestrator(project_id: int = 0) -> AuditOrchestrator:
    """Get or create the global audit orchestrator instance."""
    global _audit_orchestrator
    if _audit_orchestrator is None or _audit_orchestrator.project_id != project_id:
        _audit_orchestrator = AuditOrchestrator(project_id=project_id)
    return _audit_orchestrator

